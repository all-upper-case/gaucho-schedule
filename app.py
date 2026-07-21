from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import sqlite3
import sys
import threading
import uuid
import webbrowser
from collections import Counter, defaultdict
from contextlib import closing
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from tempfile import NamedTemporaryFile

from flask import Flask, Response, flash, jsonify, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SOURCE_DIR = Path(__file__).resolve().parent
RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", SOURCE_DIR))
IS_PACKAGED = bool(getattr(sys, "frozen", False))

if os.environ.get("GAUCHO_SCHEDULE_DATA_DIR"):
    DATA_DIR = Path(os.environ["GAUCHO_SCHEDULE_DATA_DIR"])
elif IS_PACKAGED:
    local_app_data = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local"))
    DATA_DIR = local_app_data / "GauchoSchedule"
else:
    DATA_DIR = SOURCE_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = Path(os.environ.get("GAUCHO_SCHEDULE_DB", DATA_DIR / "gaucho_schedule.sqlite3"))

app = Flask(
    __name__,
    template_folder=str(RESOURCE_DIR / "templates"),
    static_folder=str(RESOURCE_DIR / "static"),
)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-local-only-change-before-hosting")

DAYS = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
DAY_WORDS = set(DAYS + ["TUES", "WEDS", "THURS"])
GLOBAL_SHIFT_OPTIONS = [
    "9:00 AM / 4:00 PM",
    "10:00 AM / 4:00 PM",
    "3:00 PM",
    "4:00 PM",
    "5:00 PM",
    "CLOSE",
    "OFF",
]
DEFAULT_ROLES = [
    (1, "SERVERS", "(ALCOHOL)"),
    (2, "SERVERS", "(FOOD ONLY)"),
    (3, "GAUCHOS", "(MEAT)"),
    (4, "SERVER", "ASSISTANT"),
    (5, "HOST", ""),
    (6, "KITCHEN", ""),
    (7, "DISHWASHER", ""),
]
DEFAULT_EMPLOYEES = [
    ("EDWARD", 1), ("BRITTANY", 1), ("HUGO", 1), ("JARECK", 1), ("ERICK", 1), ("ROMER", 1), ("DANIEL P.", 1), ("JORDAN", 1), ("SAM", 1),
    ("MIGUEL", 2), ("DANIEL", 2), ("DIANA", 2), ("MICHAEL", 2),
    ("REYNALDO", 3), ("DAVID M.", 3), ("JUAN", 3), ("ENRIQUE", 3), ("NICOLAS", 3), ("WILLIAM", 3), ("KEVIN", 3), ("JAIDER", 3), ("LUCAS", 3), ("ISABEL", 3), ("YERSON", 3), ("JOEL", 3), ("JULIO", 3),
    ("MARTIN", 4), ("KISHANA", 4), ("LEXI", 4),
    ("EMILIO", 5), ("NICOLE", 5), ("ERIKA", 5), ("JEAN", 5),
    ("ROBERTO", 6), ("SVETLANA", 6), ("JARECK", 6), ("RAFAEL", 6), ("JESUS", 6), ("MICHAEL", 6), ("CINTIA", 6), ("ANATOLE", 6), ("DANIEL", 6), ("JESSICA", 6), ("ALEXIS", 6),
    ("ROBIN", 7), ("JONH", 7), ("FRANKLIN", 7), ("JOSETH", 7),
]


@dataclass
class Role:
    id: int
    title: str
    subtitle: str
    display_order: int


@dataclass
class ScheduleEmployee:
    assignment_id: int
    employee_id: int
    name: str
    role_id: int
    display_order: int
    archived_at: str | None = None


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    return conn


def table_exists(db: sqlite3.Connection, table: str) -> bool:
    return db.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone() is not None


def ensure_column(db: sqlite3.Connection, table: str, column: str, ddl: str) -> None:
    cols = {row["name"] for row in db.execute(f"PRAGMA table_info({table})")}
    if column not in cols:
        db.execute(f"ALTER TABLE {table} ADD COLUMN {column} {ddl}")


def init_db() -> None:
    with closing(get_db()) as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS roles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                subtitle TEXT NOT NULL DEFAULT '',
                display_order INTEGER NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                role_id INTEGER NOT NULL REFERENCES roles(id),
                active INTEGER NOT NULL DEFAULT 1,
                display_order INTEGER NOT NULL DEFAULT 0,
                archived_at TEXT
            );
            CREATE TABLE IF NOT EXISTS weeks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                start_date TEXT NOT NULL UNIQUE,
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS shifts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                week_id INTEGER NOT NULL REFERENCES weeks(id) ON DELETE CASCADE,
                employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                day_index INTEGER NOT NULL CHECK(day_index BETWEEN 0 AND 6),
                label TEXT NOT NULL DEFAULT '',
                UNIQUE(week_id, employee_id, day_index)
            );
            CREATE TABLE IF NOT EXISTS shift_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                role_title TEXT NOT NULL,
                role_subtitle TEXT NOT NULL DEFAULT '',
                employee_name TEXT NOT NULL,
                day_index INTEGER NOT NULL CHECK(day_index BETWEEN 0 AND 6),
                label TEXT NOT NULL,
                source_sheet TEXT NOT NULL DEFAULT '',
                count INTEGER NOT NULL DEFAULT 1,
                UNIQUE(role_title, role_subtitle, employee_name, day_index, label, source_sheet)
            );
            CREATE TABLE IF NOT EXISTS employee_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
                role_id INTEGER NOT NULL REFERENCES roles(id),
                active INTEGER NOT NULL DEFAULT 1,
                display_order INTEGER NOT NULL DEFAULT 0,
                archived_at TEXT,
                UNIQUE(employee_id, role_id)
            );
            CREATE TABLE IF NOT EXISTS schedule_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                week_id INTEGER NOT NULL REFERENCES weeks(id) ON DELETE CASCADE,
                assignment_id INTEGER NOT NULL REFERENCES employee_assignments(id) ON DELETE CASCADE,
                day_index INTEGER NOT NULL CHECK(day_index BETWEEN 0 AND 6),
                label TEXT NOT NULL DEFAULT '',
                UNIQUE(week_id, assignment_id, day_index)
            );
            CREATE TABLE IF NOT EXISTS week_versions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                week_id INTEGER NOT NULL REFERENCES weeks(id) ON DELETE CASCADE,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                reason TEXT NOT NULL,
                snapshot_json TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS pos_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                fingerprint TEXT NOT NULL UNIQUE,
                date_min TEXT NOT NULL,
                date_max TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                imported_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS pos_job_mappings (
                pos_job TEXT PRIMARY KEY,
                role_id INTEGER REFERENCES roles(id),
                ignored INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS pos_employee_mappings (
                pos_employee TEXT PRIMARY KEY,
                employee_id INTEGER REFERENCES employees(id),
                ignored INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS pos_employee_job_mappings (
                pos_employee TEXT NOT NULL,
                pos_job TEXT NOT NULL,
                role_id INTEGER REFERENCES roles(id),
                ignored INTEGER NOT NULL DEFAULT 0,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (pos_employee, pos_job)
            );
            """
        )
        ensure_column(db, "employees", "archived_at", "TEXT")
        ensure_column(db, "weeks", "published_at", "TEXT")

        if db.execute("SELECT COUNT(*) FROM roles").fetchone()[0] == 0:
            db.executemany(
                "INSERT INTO roles (display_order, title, subtitle) VALUES (?, ?, ?)",
                [(order, title, subtitle) for order, title, subtitle in DEFAULT_ROLES],
            )
        if db.execute("SELECT COUNT(*) FROM employees").fetchone()[0] == 0:
            counters: dict[int, int] = {}
            rows = []
            for name, role_id in DEFAULT_EMPLOYEES:
                counters[role_id] = counters.get(role_id, 0) + 1
                rows.append((name, role_id, 1, counters[role_id]))
            db.executemany(
                "INSERT INTO employees (name, role_id, active, display_order) VALUES (?, ?, ?, ?)", rows
            )

        if db.execute("SELECT COUNT(*) FROM employee_assignments").fetchone()[0] == 0:
            db.execute(
                """INSERT OR IGNORE INTO employee_assignments
                   (employee_id, role_id, active, display_order, archived_at)
                   SELECT id, role_id, active, display_order, archived_at FROM employees"""
            )
        if db.execute("SELECT COUNT(*) FROM schedule_entries").fetchone()[0] == 0 and table_exists(db, "shifts"):
            db.execute(
                """INSERT OR IGNORE INTO schedule_entries (week_id, assignment_id, day_index, label)
                   SELECT s.week_id, a.id, s.day_index, s.label
                   FROM shifts s
                   JOIN employees e ON e.id = s.employee_id
                   JOIN employee_assignments a ON a.employee_id = e.id AND a.role_id = e.role_id"""
            )
        db.commit()


def monday_for(value: date | None = None) -> date:
    value = value or date.today()
    return value - timedelta(days=value.weekday())


def parse_week_start(raw: str | None) -> date:
    if not raw:
        return monday_for()
    return monday_for(datetime.strptime(raw, "%Y-%m-%d").date())


def get_or_create_week(db: sqlite3.Connection, week_start: date) -> int:
    row = db.execute("SELECT id FROM weeks WHERE start_date = ?", (week_start.isoformat(),)).fetchone()
    if row:
        return int(row["id"])
    cur = db.execute("INSERT INTO weeks (start_date) VALUES (?)", (week_start.isoformat(),))
    db.commit()
    return int(cur.lastrowid)


def week_row(db: sqlite3.Connection, week_start: date) -> sqlite3.Row:
    week_id = get_or_create_week(db, week_start)
    return db.execute("SELECT * FROM weeks WHERE id = ?", (week_id,)).fetchone()


def is_week_locked(db: sqlite3.Connection, week_id: int) -> bool:
    return db.execute("SELECT published_at FROM weeks WHERE id = ?", (week_id,)).fetchone()[0] is not None


def roles(db: sqlite3.Connection) -> list[Role]:
    return [Role(**dict(row)) for row in db.execute("SELECT * FROM roles ORDER BY display_order, id")]


def schedule_employees(db: sqlite3.Connection, archived: bool = False) -> list[ScheduleEmployee]:
    where = "a.archived_at IS NOT NULL" if archived else "a.active = 1 AND a.archived_at IS NULL AND e.active = 1"
    rows = db.execute(
        f"""SELECT a.id AS assignment_id, e.id AS employee_id, e.name, a.role_id,
                   a.display_order, a.archived_at
            FROM employee_assignments a
            JOIN employees e ON e.id = a.employee_id
            WHERE {where}
            ORDER BY a.role_id, a.display_order, e.name, a.id"""
    )
    return [ScheduleEmployee(**dict(row)) for row in rows]


def grouped_employees(db: sqlite3.Connection, archived: bool = False):
    all_roles = roles(db)
    by_role: dict[int, list[ScheduleEmployee]] = {role.id: [] for role in all_roles}
    for employee in schedule_employees(db, archived=archived):
        by_role.setdefault(employee.role_id, []).append(employee)
    return [(role, by_role.get(role.id, [])) for role in all_roles]


def assignment_lookup(db: sqlite3.Connection) -> dict[int, ScheduleEmployee]:
    return {employee.assignment_id: employee for employee in schedule_employees(db)}


def shift_map(db: sqlite3.Connection, week_id: int) -> dict[tuple[int, int], str]:
    rows = db.execute(
        "SELECT assignment_id, day_index, label FROM schedule_entries WHERE week_id = ?", (week_id,)
    ).fetchall()
    return {(int(row["assignment_id"]), int(row["day_index"])): row["label"] for row in rows}


def grouped_schedule(db: sqlite3.Connection, week_start: date):
    week = week_row(db, week_start)
    return week, grouped_employees(db), shift_map(db, int(week["id"]))


def time_to_12h(hour: int, minute: int = 0) -> str:
    hour %= 24
    return f"{hour % 12 or 12}:{minute:02d} {'AM' if hour < 12 else 'PM'}"


def format_minutes(total_minutes: int, output_format: str = "12h") -> str:
    total_minutes %= 24 * 60
    hour, minute = divmod(total_minutes, 60)
    return f"{hour:02d}:{minute:02d}" if output_format == "24h" else time_to_12h(hour, minute)


def normalize_one_time(value, output_format: str = "12h") -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    upper = raw.upper()
    if upper in {"CL", "CLOSE", "CLOSING"}:
        return "CLOSE"
    compact = re.fullmatch(r"\d{3,4}", raw)
    if compact:
        hour, minute = int(raw[:-2]), int(raw[-2:])
        if hour <= 23 and minute < 60:
            if hour <= 7:
                hour += 12
            return format_minutes(hour * 60 + minute, output_format)
    match = re.fullmatch(r"(\d{1,2})(?::(\d{1,2}))?\s*([AaPp][Mm])?", raw.replace(".", ":"))
    if match:
        hour = int(match.group(1))
        minute = int(match.group(2) or 0)
        suffix = match.group(3).upper() if match.group(3) else None
        if minute >= 60:
            return upper
        if suffix == "PM" and hour < 12:
            hour += 12
        elif suffix == "AM" and hour == 12:
            hour = 0
        elif not suffix and hour <= 7:
            hour += 12
        if hour <= 23:
            return format_minutes(hour * 60 + minute, output_format)
    return upper


def normalize_shift_label(value, output_format: str = "12h") -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, time):
        return format_minutes(value.hour * 60 + value.minute, output_format)
    if isinstance(value, datetime):
        return format_minutes(value.hour * 60 + value.minute, output_format) if value.hour or value.minute else ""
    if isinstance(value, (int, float)):
        if float(value) == 0:
            return "OFF"
        if 0 < float(value) < 1:
            return format_minutes(round(float(value) * 24 * 60), output_format)
    raw = str(value).strip()
    if not raw:
        return ""
    upper = raw.upper()
    if upper in {"0", "OFF", "OOF"}:
        return "OFF"
    if upper in {"CL", "CLOSE", "CLOSING"}:
        return "CLOSE"

    multi_parts = [part.strip() for part in re.split(r"\s*(?:/|,|\s+&\s+)\s*", raw) if part.strip()]
    normalized_parts: list[str] = []
    for part in multi_parts:
        # Older versions accepted start-end ranges. If one is pasted, keep only
        # the report time so printed schedules never imply a guaranteed end.
        start_only = re.split(r"\s*[;\-–—]\s*", part, maxsplit=1)[0].strip()
        normalized_parts.append(normalize_one_time(start_only, output_format))
    return " / ".join(normalized_parts)


def label_is_usable(label: str, include_off: bool = True) -> bool:
    if not label:
        return False
    if label == "OFF":
        return include_off
    if label == "CLOSE":
        return True
    return bool(re.search(r"\d{1,2}:\d{2}", label))


def unique_options(*groups: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for group in groups:
        for item in group:
            clean = (item or "").strip()
            if clean and clean not in seen:
                seen.add(clean)
                output.append(clean)
    return output


def previous_week_options(db: sqlite3.Connection, week_start: date) -> dict[tuple[int, int], str]:
    row = db.execute("SELECT id FROM weeks WHERE start_date = ?", ((week_start - timedelta(days=7)).isoformat(),)).fetchone()
    return shift_map(db, int(row["id"])) if row else {}


def history_options_for_cell(db: sqlite3.Connection, role: Role, employee: ScheduleEmployee, day_index: int) -> list[str]:
    rows = db.execute(
        """SELECT label, SUM(count) AS total FROM shift_history
           WHERE role_title = ? AND role_subtitle = ? AND employee_name = ? AND day_index = ? AND label != 'OFF'
           GROUP BY label ORDER BY total DESC, label LIMIT 12""",
        (role.title, role.subtitle, employee.name.upper(), day_index),
    ).fetchall()
    return [row["label"] for row in rows]


def role_options_for_cell(db: sqlite3.Connection, role: Role, day_index: int) -> list[str]:
    rows = db.execute(
        """SELECT label, SUM(count) AS total FROM shift_history
           WHERE role_title = ? AND role_subtitle = ? AND day_index = ? AND label != 'OFF'
           GROUP BY label ORDER BY total DESC, label LIMIT 10""",
        (role.title, role.subtitle, day_index),
    ).fetchall()
    return [row["label"] for row in rows]


def global_options_for_cell(db: sqlite3.Connection, day_index: int) -> list[str]:
    rows = db.execute(
        """SELECT label, SUM(count) AS total FROM shift_history
           WHERE day_index = ? AND label != 'OFF'
           GROUP BY label ORDER BY total DESC, label LIMIT 10""",
        (day_index,),
    ).fetchall()
    return [row["label"] for row in rows]


def best_label_for_cell(db: sqlite3.Connection, role: Role, employee: ScheduleEmployee, day_index: int) -> str:
    sources = [
        history_options_for_cell(db, role, employee, day_index),
        role_options_for_cell(db, role, day_index),
        global_options_for_cell(db, day_index),
        GLOBAL_SHIFT_OPTIONS,
    ]
    for source in sources:
        for label in source:
            if label and label != "OFF":
                return label
    return "OFF"


def build_suggestions(db: sqlite3.Connection, grouped, week_start: date) -> dict[tuple[int, int], list[str]]:
    previous = previous_week_options(db, week_start)
    suggestions: dict[tuple[int, int], list[str]] = {}
    for role, employees_for_role in grouped:
        for employee in employees_for_role:
            for day_index in range(7):
                prev = previous.get((employee.assignment_id, day_index))
                suggestions[(employee.assignment_id, day_index)] = unique_options(
                    [prev] if prev else [],
                    history_options_for_cell(db, role, employee, day_index),
                    role_options_for_cell(db, role, day_index),
                    global_options_for_cell(db, day_index),
                    GLOBAL_SHIFT_OPTIONS,
                )[:20]
    return suggestions


def save_shift_value(db: sqlite3.Connection, week_id: int, employee: ScheduleEmployee, role: Role, day_index: int, raw_label: str) -> str:
    label = normalize_shift_label(raw_label or "", "12h")
    db.execute(
        """INSERT INTO schedule_entries (week_id, assignment_id, day_index, label)
           VALUES (?, ?, ?, ?)
           ON CONFLICT(week_id, assignment_id, day_index) DO UPDATE SET label = excluded.label""",
        (week_id, employee.assignment_id, day_index, label),
    )
    if label_is_usable(label):
        db.execute(
            """INSERT INTO shift_history
               (role_title, role_subtitle, employee_name, day_index, label, source_sheet, count)
               VALUES (?, ?, ?, ?, ?, 'manual-app-entry', 1)
               ON CONFLICT(role_title, role_subtitle, employee_name, day_index, label, source_sheet)
               DO UPDATE SET count = count + 1""",
            (role.title, role.subtitle, employee.name.upper(), day_index, label),
        )
    return label


def week_dates(week_start: date) -> list[date]:
    return [week_start + timedelta(days=index) for index in range(7)]


def short_date_label(value: date) -> str:
    return f"{DAYS[value.weekday()]} {value.month}/{value.day}/{str(value.year)[-2:]}"


def balanced_print_pages(grouped: list[tuple[Role, list[ScheduleEmployee]]]) -> list[list[tuple[Role, list[ScheduleEmployee]]]]:
    """Split complete role groups at the boundary closest to half the printed rows."""
    if len(grouped) < 2:
        return [grouped]
    weights = [len(employees) + 2 for _, employees in grouped]
    total = sum(weights)
    running = 0
    best_cut = 1
    best_difference = total
    for index, weight in enumerate(weights[:-1], start=1):
        running += weight
        difference = abs(total - (2 * running))
        if difference < best_difference:
            best_cut = index
            best_difference = difference
    return [grouped[:best_cut], grouped[best_cut:]]


def surrounding_weeks(center: date, radius: int = 4):
    return [center + timedelta(days=7 * offset) for offset in range(-radius, radius + 1)]


def create_week_snapshot(db: sqlite3.Connection, week_id: int, reason: str) -> int:
    week = dict(db.execute("SELECT * FROM weeks WHERE id = ?", (week_id,)).fetchone())
    entries = [dict(row) for row in db.execute(
        """SELECT se.assignment_id, e.name, r.title AS role, r.subtitle, se.day_index, se.label
           FROM schedule_entries se
           JOIN employee_assignments a ON a.id = se.assignment_id
           JOIN employees e ON e.id = a.employee_id
           JOIN roles r ON r.id = a.role_id
           WHERE se.week_id = ? ORDER BY r.display_order, a.display_order, se.day_index""", (week_id,)
    )]
    cur = db.execute(
        "INSERT INTO week_versions (week_id, reason, snapshot_json) VALUES (?, ?, ?)",
        (week_id, reason, json.dumps({"week": week, "entries": entries}, indent=2)),
    )
    return int(cur.lastrowid)


def is_header_row(row) -> bool:
    return [str(row[index].value or "").strip().upper() for index in range(2, 9)] == DAYS


def import_history_from_workbook(path: Path) -> tuple[int, int, int]:
    book = load_workbook(path, data_only=True, read_only=True, keep_vba=True)
    rows_seen = 0
    shift_count = 0
    sheet_count = 0
    inserts = []
    for worksheet in book.worksheets:
        if worksheet.title.upper().startswith("SHEET"):
            continue
        sheet_count += 1
        current_role: tuple[str, str] | None = None
        sheet_rows = list(worksheet.iter_rows(min_row=1, max_row=130, min_col=1, max_col=9))
        for index, row in enumerate(sheet_rows):
            if is_header_row(row):
                title = str(row[0].value or "").strip().upper()
                subtitle = str(sheet_rows[index + 1][0].value or "").strip().upper() if index + 1 < len(sheet_rows) else ""
                current_role = (title, subtitle)
                continue
            if not current_role:
                continue
            name = str(row[0].value or "").strip().upper()
            if not name or name in DAY_WORDS or "SIGNATURE" in name or name.startswith("("):
                continue
            labels = [normalize_shift_label(cell.value, "12h") for cell in row[2:9]]
            if not any(label_is_usable(label, include_off=False) for label in labels):
                continue
            rows_seen += 1
            for day_index, label in enumerate(labels):
                if label_is_usable(label):
                    inserts.append((current_role[0], current_role[1], name, day_index, label, worksheet.title))
                    shift_count += 1
    book.close()
    with closing(get_db()) as db:
        db.execute("DELETE FROM shift_history WHERE source_sheet != 'manual-app-entry' AND source_sheet NOT LIKE 'POS:%'")
        db.executemany(
            """INSERT INTO shift_history
               (role_title, role_subtitle, employee_name, day_index, label, source_sheet, count)
               VALUES (?, ?, ?, ?, ?, ?, 1)
               ON CONFLICT(role_title, role_subtitle, employee_name, day_index, label, source_sheet)
               DO UPDATE SET count = count + 1""", inserts
        )
        db.commit()
    return sheet_count, rows_seen, shift_count


POS_REQUIRED_COLUMNS = {"Employee", "Job", "Date", "Time In", "Time Out", "Total Hours"}


def parse_pos_date(raw: str) -> date:
    value = (raw or "").strip()
    for fmt in ("%d-%b-%y", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unrecognized POS date: {value}")


def parse_pos_time(raw: str) -> int | None:
    value = (raw or "").strip()
    if not value:
        return None
    for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M"):
        try:
            parsed = datetime.strptime(value, fmt)
            return parsed.hour * 60 + parsed.minute
        except ValueError:
            pass
    raise ValueError(f"Unrecognized POS time: {value}")


def round_report_time(total_minutes: int, step: int = 15) -> str:
    rounded = int((total_minutes + step / 2) // step * step) % (24 * 60)
    return format_minutes(rounded, "12h")


def read_pos_csv(path: Path) -> tuple[list[dict], list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        columns = [str(column or "").strip() for column in (reader.fieldnames or [])]
        missing = sorted(POS_REQUIRED_COLUMNS - set(columns))
        if missing:
            raise ValueError(f"Missing required column(s): {', '.join(missing)}")
        rows = []
        errors = []
        for line_number, source in enumerate(reader, start=2):
            clean = {key: (source.get(key) or "").strip() for key in POS_REQUIRED_COLUMNS}
            if not clean["Employee"] or not clean["Job"] or not clean["Date"] or not clean["Time In"]:
                errors.append(f"Line {line_number}: employee, job, date, or Time In is blank")
                continue
            try:
                clean["_date"] = parse_pos_date(clean["Date"])
                clean["_in"] = parse_pos_time(clean["Time In"])
                clean["_out"] = parse_pos_time(clean["Time Out"])
            except ValueError as exc:
                errors.append(f"Line {line_number}: {exc}")
                continue
            clean["_line"] = line_number
            rows.append(clean)
    if not rows:
        raise ValueError("No usable time entries were found in the CSV.")
    return rows, errors


def classify_pos_starts(rows: list[dict]) -> tuple[list[dict], dict[str, int]]:
    by_employee_day: dict[tuple[str, date], list[dict]] = defaultdict(list)
    for row in rows:
        by_employee_day[(row["Employee"], row["_date"])].append(row)

    learned: dict[tuple[str, str, date], list[int]] = defaultdict(list)
    stats = Counter()
    stats["employee_days"] = len(by_employee_day)
    stats["multi_entry_days"] = sum(len(day_rows) > 1 for day_rows in by_employee_day.values())

    for (_, _), day_rows in by_employee_day.items():
        ordered = sorted(day_rows, key=lambda row: (int(row["_in"]), int(row["_line"])))
        previous = None
        for row in ordered:
            learn = previous is None
            classification = "first_start" if previous is None else "ambiguous"
            if previous is not None:
                previous_out = previous["_out"]
                if previous_out is None:
                    learn = False
                    classification = "ambiguous"
                else:
                    adjusted_out = int(previous_out)
                    if adjusted_out < int(previous["_in"]):
                        adjusted_out += 24 * 60
                    current_in = int(row["_in"])
                    if current_in < int(previous["_in"]) and adjusted_out >= 24 * 60:
                        current_in += 24 * 60
                    gap = current_in - adjusted_out
                    same_job = row["Job"] == previous["Job"]
                    if same_job and 0 <= gap <= 60:
                        learn = False
                        classification = "meal_or_reclock"
                    elif not same_job and -5 <= gap <= 30:
                        learn = False
                        classification = "job_transfer"
                    elif gap >= 90:
                        learn = True
                        classification = "split_start"
                    else:
                        learn = False
                        classification = "ambiguous"
            stats[classification] += 1
            if learn:
                learned[(row["Employee"], row["Job"], row["_date"])].append(int(row["_in"]))
            previous = row

    patterns = []
    for (employee, job, work_date), starts in sorted(learned.items()):
        labels = []
        for start in sorted(starts):
            label = round_report_time(start)
            if label not in labels:
                labels.append(label)
        if labels:
            patterns.append({
                "employee": employee,
                "job": job,
                "date": work_date,
                "day_index": work_date.weekday(),
                "label": " / ".join(labels),
            })
    stats["patterns"] = len(patterns)
    return patterns, dict(stats)


def pos_file_fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalized_person_name(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", " ", (value or "").upper()).strip()


def pos_name_aliases(pos_name: str) -> set[str]:
    raw = (pos_name or "").strip()
    if "," in raw:
        last, given = [part.strip() for part in raw.split(",", 1)]
    else:
        pieces = raw.split()
        given, last = (pieces[0], " ".join(pieces[1:])) if pieces else (raw, "")
    first = given.split()[0] if given else ""
    last_initial = last[:1]
    candidates = {raw, f"{given} {last}", first, f"{first} {last_initial}", f"{first} {last_initial}."}
    return {normalized_person_name(candidate) for candidate in candidates if candidate.strip()}


def default_schedule_name(pos_name: str) -> str:
    raw = (pos_name or "").strip()
    if "," in raw:
        _, given = [part.strip() for part in raw.split(",", 1)]
        return (given.split()[0] if given else raw).upper()
    return (raw.split()[0] if raw else raw).upper()


def mapping_employee_options(db: sqlite3.Connection) -> list[dict]:
    rows = db.execute(
        """SELECT e.id, e.name, e.active,
                  GROUP_CONCAT(TRIM(r.title || ' ' || r.subtitle), ', ') AS role_names
           FROM employees e
           LEFT JOIN employee_assignments a ON a.employee_id = e.id AND a.archived_at IS NULL
           LEFT JOIN roles r ON r.id = a.role_id
           GROUP BY e.id ORDER BY e.name, e.id"""
    ).fetchall()
    return [dict(row) for row in rows]


def suggested_employee_id(db: sqlite3.Connection, pos_name: str, options: list[dict]) -> int | None:
    saved = db.execute("SELECT employee_id, ignored FROM pos_employee_mappings WHERE pos_employee = ?", (pos_name,)).fetchone()
    if saved:
        return None if saved["ignored"] else saved["employee_id"]
    aliases = pos_name_aliases(pos_name)
    exact = [option["id"] for option in options if normalized_person_name(option["name"]) in aliases]
    return int(exact[0]) if len(exact) == 1 else None


def role_heuristics(all_roles: list[Role]) -> dict[str, int]:
    result = {}
    for role in all_roles:
        title = normalized_person_name(role.title)
        subtitle = normalized_person_name(role.subtitle)
        if title == "KITCHEN":
            result["Kitchen Aid"] = role.id
        elif title == "DISHWASHER":
            result["Dish Washer"] = role.id
        elif title == "HOST":
            result["Host-Hostess"] = role.id
        elif title == "SERVER" and subtitle == "ASSISTANT":
            result["Server Assistant"] = role.id
    return result


def ensure_employee_assignment(db: sqlite3.Connection, employee_id: int, role_id: int) -> int:
    existing = db.execute("SELECT id FROM employee_assignments WHERE employee_id = ? AND role_id = ?", (employee_id, role_id)).fetchone()
    if existing:
        db.execute("UPDATE employee_assignments SET active = 1, archived_at = NULL WHERE id = ?", (existing["id"],))
        db.execute("UPDATE employees SET active = 1, archived_at = NULL WHERE id = ?", (employee_id,))
        return int(existing["id"])
    max_order = db.execute("SELECT COALESCE(MAX(display_order), 0) FROM employee_assignments WHERE role_id = ? AND archived_at IS NULL", (role_id,)).fetchone()[0]
    cur = db.execute("INSERT INTO employee_assignments (employee_id, role_id, active, display_order) VALUES (?, ?, 1, ?)", (employee_id, role_id, int(max_order) + 1))
    return int(cur.lastrowid)


def coverage_data(grouped, shifts):
    rows = []
    for role, employees_for_role in grouped:
        counts = []
        names = []
        for day_index in range(7):
            scheduled = [
                employee.name
                for employee in employees_for_role
                if shifts.get((employee.assignment_id, day_index), "").strip() not in {"", "OFF"}
            ]
            counts.append(len(scheduled))
            names.append(scheduled)
        rows.append((role, counts, names))
    return rows


@app.before_request
def ensure_db() -> None:
    init_db()


@app.route("/")
def index():
    current = monday_for()
    with closing(get_db()) as db:
        week_rows = db.execute("SELECT * FROM weeks ORDER BY start_date DESC LIMIT 20").fetchall()
        history_rows = db.execute("SELECT COUNT(*) FROM shift_history").fetchone()[0]
        history_times = db.execute("SELECT COUNT(*) FROM shift_history WHERE label != 'OFF'").fetchone()[0]
    return render_template(
        "index.html", weeks=week_rows, current_week=current, week_nav=surrounding_weeks(current),
        seven_days=timedelta(days=6), history_rows=history_rows, history_times=history_times,
        data_path=str(DB_PATH), packaged=IS_PACKAGED,
    )


@app.route("/week", methods=["POST"])
def go_to_week():
    return redirect(url_for("edit_week", week_start=parse_week_start(request.form.get("week_start")).isoformat()))


@app.route("/week/<week_start>")
def edit_week(week_start: str):
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        week, grouped, shifts = grouped_schedule(db, start)
        suggestions = build_suggestions(db, grouped, start)
        versions = db.execute("SELECT id, created_at, reason FROM week_versions WHERE week_id = ? ORDER BY id DESC", (week["id"],)).fetchall()
    return render_template(
        "schedule.html", week=week, week_start=start, week_end=start + timedelta(days=6),
        dates=week_dates(start), days=DAYS, grouped=grouped, shifts=shifts,
        suggestions=suggestions, versions=versions,
    )


@app.route("/api/week/<week_start>/shift", methods=["POST"])
def api_save_shift(week_start: str):
    data = request.get_json(force=True)
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        week = week_row(db, start)
        if week["published_at"]:
            return jsonify(ok=False, error="This schedule is published and locked."), 409
        employee = assignment_lookup(db).get(int(data["assignment_id"]))
        if not employee:
            return jsonify(ok=False, error="Assignment not found"), 404
        role = {role.id: role for role in roles(db)}[employee.role_id]
        label = save_shift_value(db, int(week["id"]), employee, role, int(data["day_index"]), data.get("label", ""))
        db.commit()
    return jsonify(ok=True, label=label)


@app.route("/api/week/<week_start>/fill-best", methods=["POST"])
def api_fill_best(week_start: str):
    start = parse_week_start(week_start)
    payload = {}
    with closing(get_db()) as db:
        week = week_row(db, start)
        if week["published_at"]:
            return jsonify(ok=False, error="This schedule is published and locked."), 409
        for role, employees_for_role in grouped_employees(db):
            for employee in employees_for_role:
                for day_index in range(7):
                    label = save_shift_value(db, int(week["id"]), employee, role, day_index, best_label_for_cell(db, role, employee, day_index))
                    payload[f"shift_{employee.assignment_id}_{day_index}"] = label
        db.commit()
    return jsonify(ok=True, shifts=payload)


@app.route("/api/week/<week_start>/fill-off", methods=["POST"])
def api_fill_off(week_start: str):
    start = parse_week_start(week_start)
    payload = {}
    with closing(get_db()) as db:
        week, grouped, shifts = grouped_schedule(db, start)
        if week["published_at"]:
            return jsonify(ok=False, error="This schedule is published and locked."), 409
        for role, employees_for_role in grouped:
            for employee in employees_for_role:
                for day_index in range(7):
                    if not shifts.get((employee.assignment_id, day_index), "").strip():
                        label = save_shift_value(db, int(week["id"]), employee, role, day_index, "OFF")
                        payload[f"shift_{employee.assignment_id}_{day_index}"] = label
        db.commit()
    return jsonify(ok=True, shifts=payload)


@app.route("/api/week/<week_start>/copy-previous", methods=["POST"])
def api_copy_previous(week_start: str):
    start = parse_week_start(week_start)
    payload = {}
    with closing(get_db()) as db:
        week = week_row(db, start)
        if week["published_at"]:
            return jsonify(ok=False, error="This schedule is published and locked."), 409
        previous = previous_week_options(db, start)
        if db.execute("SELECT COUNT(*) FROM schedule_entries WHERE week_id = ? AND label != ''", (week["id"],)).fetchone()[0]:
            create_week_snapshot(db, int(week["id"]), "Before replacing with previous week")
        for role, employees_for_role in grouped_employees(db):
            for employee in employees_for_role:
                for day_index in range(7):
                    label = save_shift_value(db, int(week["id"]), employee, role, day_index, previous.get((employee.assignment_id, day_index), ""))
                    payload[f"shift_{employee.assignment_id}_{day_index}"] = label
        db.commit()
    return jsonify(ok=True, shifts=payload)


@app.route("/week/<week_start>/save", methods=["POST"])
def save_week(week_start: str):
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        week = week_row(db, start)
        if week["published_at"]:
            flash("This schedule is published and locked. Reopen it before making corrections.")
            return redirect(url_for("edit_week", week_start=start.isoformat()))
        for role, employees_for_role in grouped_employees(db):
            for employee in employees_for_role:
                for day_index in range(7):
                    save_shift_value(db, int(week["id"]), employee, role, day_index, request.form.get(f"shift_{employee.assignment_id}_{day_index}", ""))
        db.commit()
    flash("Schedule saved.")
    return redirect(url_for("edit_week", week_start=start.isoformat()))


@app.route("/week/<week_start>/publish", methods=["POST"])
def publish_week(week_start: str):
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        week = week_row(db, start)
        if not week["published_at"]:
            for role, employees_for_role in grouped_employees(db):
                for employee in employees_for_role:
                    for day_index in range(7):
                        field = f"shift_{employee.assignment_id}_{day_index}"
                        if field in request.form:
                            save_shift_value(db, int(week["id"]), employee, role, day_index, request.form[field])
            create_week_snapshot(db, int(week["id"]), "Published schedule")
            db.execute("UPDATE weeks SET published_at = CURRENT_TIMESTAMP WHERE id = ?", (week["id"],))
            db.commit()
            flash("Schedule published, snapshotted, and locked.")
    return redirect(url_for("edit_week", week_start=start.isoformat()))


@app.route("/week/<week_start>/reopen", methods=["POST"])
def reopen_week(week_start: str):
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        week = week_row(db, start)
        if week["published_at"]:
            create_week_snapshot(db, int(week["id"]), "Before reopening published schedule")
            db.execute("UPDATE weeks SET published_at = NULL WHERE id = ?", (week["id"],))
            db.commit()
            flash("Schedule reopened. The published copy remains in version history.")
    return redirect(url_for("edit_week", week_start=start.isoformat()))


@app.route("/week/<week_start>/copy-next", methods=["POST"])
def copy_next_week(week_start: str):
    start = parse_week_start(week_start)
    next_start = start + timedelta(days=7)
    with closing(get_db()) as db:
        source = week_row(db, start)
        target = week_row(db, next_start)
        existing = db.execute("SELECT COUNT(*) FROM schedule_entries WHERE week_id = ? AND label != ''", (target["id"],)).fetchone()[0]
        if existing:
            flash("The next week already contains schedule entries, so it was not overwritten.")
            return redirect(url_for("edit_week", week_start=next_start.isoformat()))
        if not source["published_at"]:
            for role, employees_for_role in grouped_employees(db):
                for employee in employees_for_role:
                    for day_index in range(7):
                        field = f"shift_{employee.assignment_id}_{day_index}"
                        if field in request.form:
                            save_shift_value(db, int(source["id"]), employee, role, day_index, request.form[field])
            create_week_snapshot(db, int(source["id"]), "Published when next week was created")
            db.execute("UPDATE weeks SET published_at = CURRENT_TIMESTAMP WHERE id = ?", (source["id"],))
        db.execute(
            """INSERT INTO schedule_entries (week_id, assignment_id, day_index, label)
               SELECT ?, assignment_id, day_index, label FROM schedule_entries WHERE week_id = ?""",
            (target["id"], source["id"]),
        )
        db.commit()
    flash("Current week published and preserved; next week created from it.")
    return redirect(url_for("edit_week", week_start=next_start.isoformat()))


@app.route("/week/<week_start>/versions")
def week_versions(week_start: str):
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        week = week_row(db, start)
        versions = db.execute("SELECT id, created_at, reason FROM week_versions WHERE week_id = ? ORDER BY id DESC", (week["id"],)).fetchall()
    return render_template("versions.html", week_start=start, versions=versions)


@app.route("/week/<week_start>/versions/<int:version_id>.json")
def download_week_version(week_start: str, version_id: int):
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        week = week_row(db, start)
        row = db.execute("SELECT snapshot_json FROM week_versions WHERE id = ? AND week_id = ?", (version_id, week["id"])).fetchone()
    if not row:
        return "Version not found", 404
    return Response(row["snapshot_json"], mimetype="application/json", headers={"Content-Disposition": f"attachment; filename=gaucho-schedule-{start.isoformat()}-version-{version_id}.json"})


@app.route("/coverage/<week_start>")
def coverage(week_start: str):
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        _, grouped, shifts = grouped_schedule(db, start)
    coverage_rows = coverage_data(grouped, shifts)
    return render_template("coverage.html", week_start=start, dates=week_dates(start), days=DAYS, coverage_rows=coverage_rows)


@app.route("/print/<week_start>")
def print_week(week_start: str):
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        _, grouped, shifts = grouped_schedule(db, start)
    dates = week_dates(start)
    return render_template(
        "print.html",
        week_start=start,
        week_end=start + timedelta(days=6),
        dates=dates,
        date_labels=[short_date_label(value) for value in dates],
        print_pages=balanced_print_pages(grouped),
        shifts=shifts,
    )


@app.route("/preferences")
def preferences():
    return render_template("preferences.html")


@app.route("/import-history", methods=["GET", "POST"])
def import_history():
    if request.method == "POST":
        upload = request.files.get("schedule_workbook")
        if not upload or not upload.filename:
            flash("Choose the schedule workbook first.")
            return redirect(url_for("import_history"))
        suffix = Path(upload.filename).suffix or ".xlsx"
        with NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            upload.save(temp.name)
            temp_path = Path(temp.name)
        try:
            sheet_count, row_count, shift_count = import_history_from_workbook(temp_path)
            flash(f"Imported {shift_count} shifts from {row_count} employee rows across {sheet_count} sheets.")
        finally:
            temp_path.unlink(missing_ok=True)
        return redirect(url_for("import_history"))
    with closing(get_db()) as db:
        recent_pos_imports = db.execute("SELECT * FROM pos_imports ORDER BY id DESC LIMIT 12").fetchall()
        pos_pattern_count = db.execute("SELECT COUNT(*) FROM shift_history WHERE source_sheet LIKE 'POS:%'").fetchone()[0]
    return render_template("import_history.html", recent_pos_imports=recent_pos_imports, pos_pattern_count=pos_pattern_count)


@app.route("/import-pos/preview", methods=["POST"])
def preview_pos_import():
    upload = request.files.get("pos_csv")
    if not upload or not upload.filename:
        flash("Choose the POS time-entry CSV first.")
        return redirect(url_for("import_history"))
    if Path(upload.filename).suffix.lower() != ".csv":
        flash("The POS export must be a CSV file.")
        return redirect(url_for("import_history"))

    stage_dir = DATA_DIR / "staged_pos_imports"
    stage_dir.mkdir(parents=True, exist_ok=True)
    token = uuid.uuid4().hex
    staged_path = stage_dir / f"{token}.csv"
    upload.save(staged_path)
    fingerprint = pos_file_fingerprint(staged_path)

    try:
        rows, row_errors = read_pos_csv(staged_path)
        patterns, classification = classify_pos_starts(rows)
    except ValueError as exc:
        staged_path.unlink(missing_ok=True)
        flash(str(exc))
        return redirect(url_for("import_history"))

    dates = [row["_date"] for row in rows]
    jobs = sorted({row["Job"] for row in rows})
    entries_by_employee_job = Counter((row["Employee"], row["Job"]) for row in rows)
    pos_employees = sorted({row["Employee"] for row in rows})

    with closing(get_db()) as db:
        duplicate = db.execute("SELECT * FROM pos_imports WHERE fingerprint = ?", (fingerprint,)).fetchone()
        if duplicate:
            staged_path.unlink(missing_ok=True)
            flash(f"This exact file was already imported on {duplicate['imported_at']} as {duplicate['filename']}.")
            return redirect(url_for("import_history"))

        all_roles = roles(db)
        employee_options = mapping_employee_options(db)
        employee_by_id = {int(option["id"]): option for option in employee_options}
        employee_roles = defaultdict(set)
        for row in db.execute("SELECT employee_id, role_id FROM employee_assignments WHERE archived_at IS NULL"):
            employee_roles[int(row["employee_id"])].add(int(row["role_id"]))
        heuristics = role_heuristics(all_roles)
        saved_job_rows = {row["pos_job"]: row for row in db.execute("SELECT * FROM pos_job_mappings")}
        saved_specific = {(row["pos_employee"], row["pos_job"]): row for row in db.execute("SELECT * FROM pos_employee_job_mappings")}

        job_rows = []
        for job in jobs:
            saved = saved_job_rows.get(job)
            if saved:
                default_value = "ignore" if saved["ignored"] else f"role:{saved['role_id']}"
            elif job == "MIT":
                default_value = "ignore"
            elif job in heuristics:
                default_value = f"role:{heuristics[job]}"
            else:
                default_value = ""
            job_rows.append({"job": job, "count": sum(1 for row in rows if row["Job"] == job), "default": default_value})

        employee_rows = []
        for pos_employee in pos_employees:
            saved_employee = db.execute("SELECT * FROM pos_employee_mappings WHERE pos_employee = ?", (pos_employee,)).fetchone()
            suggested_id = suggested_employee_id(db, pos_employee, employee_options)
            if saved_employee and saved_employee["ignored"]:
                employee_choice = "ignore"
            elif saved_employee and saved_employee["employee_id"]:
                employee_choice = f"existing:{saved_employee['employee_id']}"
            elif suggested_id:
                employee_choice = f"existing:{suggested_id}"
            else:
                employee_choice = "create"
            selected_employee_id = int(employee_choice.split(":", 1)[1]) if employee_choice.startswith("existing:") else None

            pair_rows = []
            employee_jobs = sorted({row["Job"] for row in rows if row["Employee"] == pos_employee})
            for pos_job in employee_jobs:
                specific = saved_specific.get((pos_employee, pos_job))
                if specific:
                    role_value = "ignore" if specific["ignored"] else f"role:{specific['role_id']}"
                else:
                    role_value = ""
                    candidate = heuristics.get(pos_job)
                    if candidate:
                        role_value = f"role:{candidate}"
                    elif selected_employee_id:
                        existing_role_ids = employee_roles.get(selected_employee_id, set())
                        server_role_ids = {
                            role.id for role in all_roles
                            if normalized_person_name(role.title) in {"SERVER", "SERVERS", "GAUCHOS"}
                        }
                        candidates = existing_role_ids & server_role_ids if pos_job == "Server" else existing_role_ids
                        if len(candidates) == 1:
                            role_value = f"role:{next(iter(candidates))}"
                    saved_job = saved_job_rows.get(pos_job)
                    if not role_value and saved_job:
                        role_value = "ignore" if saved_job["ignored"] else f"role:{saved_job['role_id']}"
                    matching_job_row = next(job_row for job_row in job_rows if job_row["job"] == pos_job)
                    if not role_value:
                        role_value = matching_job_row["default"]
                pair_rows.append({"job": pos_job, "count": entries_by_employee_job[(pos_employee, pos_job)], "role_value": role_value})
            if employee_choice == "create" and pair_rows and all(pair["role_value"] == "ignore" for pair in pair_rows):
                employee_choice = "ignore"
            employee_rows.append({
                "pos_name": pos_employee,
                "employee_choice": employee_choice,
                "new_name": default_schedule_name(pos_employee),
                "pairs": pair_rows,
            })

    return render_template(
        "pos_import_preview.html",
        token=token,
        original_filename=Path(upload.filename).name,
        fingerprint=fingerprint,
        row_count=len(rows),
        date_min=min(dates),
        date_max=max(dates),
        row_errors=row_errors,
        classification=classification,
        pattern_count=len(patterns),
        job_rows=job_rows,
        employee_rows=employee_rows,
        employee_options=employee_options,
        roles=all_roles,
    )


@app.route("/import-pos/confirm/<token>", methods=["POST"])
def confirm_pos_import(token: str):
    if not re.fullmatch(r"[0-9a-f]{32}", token):
        return "Invalid import token", 400
    staged_path = DATA_DIR / "staged_pos_imports" / f"{token}.csv"
    if not staged_path.exists():
        flash("That staged import no longer exists. Upload the CSV again.")
        return redirect(url_for("import_history"))

    try:
        rows, row_errors = read_pos_csv(staged_path)
        patterns, classification = classify_pos_starts(rows)
    except ValueError as exc:
        flash(str(exc))
        return redirect(url_for("import_history"))
    fingerprint = pos_file_fingerprint(staged_path)
    pos_employees = sorted({row["Employee"] for row in rows})
    jobs = sorted({row["Job"] for row in rows})
    pairs_by_employee = {
        employee: sorted({row["Job"] for row in rows if row["Employee"] == employee})
        for employee in pos_employees
    }
    filename = Path(request.form.get("original_filename") or "POS-time-entries.csv").name

    with closing(get_db()) as db:
        if db.execute("SELECT 1 FROM pos_imports WHERE fingerprint = ?", (fingerprint,)).fetchone():
            staged_path.unlink(missing_ok=True)
            flash("This exact POS export has already been imported.")
            return redirect(url_for("import_history"))
        valid_roles = {role.id: role for role in roles(db)}
        valid_employee_ids = {int(row["id"]) for row in db.execute("SELECT id FROM employees")}

        for job_index, pos_job in enumerate(jobs):
            choice = request.form.get(f"job_default_{job_index}", "")
            if choice == "ignore":
                db.execute("""INSERT INTO pos_job_mappings (pos_job, role_id, ignored, updated_at) VALUES (?, NULL, 1, CURRENT_TIMESTAMP)
                              ON CONFLICT(pos_job) DO UPDATE SET role_id = NULL, ignored = 1, updated_at = CURRENT_TIMESTAMP""", (pos_job,))
            elif choice.startswith("role:") and int(choice.split(":", 1)[1]) in valid_roles:
                role_id = int(choice.split(":", 1)[1])
                db.execute("""INSERT INTO pos_job_mappings (pos_job, role_id, ignored, updated_at) VALUES (?, ?, 0, CURRENT_TIMESTAMP)
                              ON CONFLICT(pos_job) DO UPDATE SET role_id = excluded.role_id, ignored = 0, updated_at = CURRENT_TIMESTAMP""", (pos_job, role_id))

        employee_id_by_pos: dict[str, int] = {}
        pattern_mapping: dict[tuple[str, str], tuple[int, int]] = {}
        created_employees = 0
        assignments_added = 0

        for employee_index, pos_employee in enumerate(pos_employees):
            action = request.form.get(f"employee_choice_{employee_index}", "ignore")
            pair_choices = []
            for pair_index, pos_job in enumerate(pairs_by_employee[pos_employee]):
                role_choice = request.form.get(f"role_choice_{employee_index}_{pair_index}", "ignore")
                if role_choice.startswith("role:") and int(role_choice.split(":", 1)[1]) in valid_roles:
                    pair_choices.append((pos_job, int(role_choice.split(":", 1)[1])))
                else:
                    pair_choices.append((pos_job, None))

            if action == "ignore":
                db.execute("""INSERT INTO pos_employee_mappings (pos_employee, employee_id, ignored, updated_at) VALUES (?, NULL, 1, CURRENT_TIMESTAMP)
                              ON CONFLICT(pos_employee) DO UPDATE SET employee_id = NULL, ignored = 1, updated_at = CURRENT_TIMESTAMP""", (pos_employee,))
                continue
            if action.startswith("existing:"):
                employee_id = int(action.split(":", 1)[1])
                if employee_id not in valid_employee_ids:
                    db.rollback()
                    return f"Invalid employee mapping for {pos_employee}", 400
            elif action == "create":
                mapped_roles = [role_id for _, role_id in pair_choices if role_id]
                if not mapped_roles:
                    db.execute("""INSERT INTO pos_employee_mappings (pos_employee, employee_id, ignored, updated_at) VALUES (?, NULL, 1, CURRENT_TIMESTAMP)
                                  ON CONFLICT(pos_employee) DO UPDATE SET employee_id = NULL, ignored = 1, updated_at = CURRENT_TIMESTAMP""", (pos_employee,))
                    continue
                schedule_name = (request.form.get(f"new_name_{employee_index}") or default_schedule_name(pos_employee)).strip().upper()
                if not schedule_name:
                    db.rollback()
                    return f"Enter a schedule name for {pos_employee}", 400
                primary_role = mapped_roles[0]
                max_order = db.execute("SELECT COALESCE(MAX(display_order), 0) FROM employee_assignments WHERE role_id = ? AND archived_at IS NULL", (primary_role,)).fetchone()[0]
                cur = db.execute("INSERT INTO employees (name, role_id, active, display_order) VALUES (?, ?, 1, ?)", (schedule_name, primary_role, int(max_order) + 1))
                employee_id = int(cur.lastrowid)
                valid_employee_ids.add(employee_id)
                created_employees += 1
            else:
                db.rollback()
                return f"Invalid employee action for {pos_employee}", 400

            employee_id_by_pos[pos_employee] = employee_id
            db.execute("""INSERT INTO pos_employee_mappings (pos_employee, employee_id, ignored, updated_at) VALUES (?, ?, 0, CURRENT_TIMESTAMP)
                          ON CONFLICT(pos_employee) DO UPDATE SET employee_id = excluded.employee_id, ignored = 0, updated_at = CURRENT_TIMESTAMP""", (pos_employee, employee_id))
            for pos_job, role_id in pair_choices:
                if role_id:
                    before = db.execute("SELECT 1 FROM employee_assignments WHERE employee_id = ? AND role_id = ?", (employee_id, role_id)).fetchone()
                    ensure_employee_assignment(db, employee_id, role_id)
                    assignments_added += 0 if before else 1
                    pattern_mapping[(pos_employee, pos_job)] = (employee_id, role_id)
                    db.execute("""INSERT INTO pos_employee_job_mappings (pos_employee, pos_job, role_id, ignored, updated_at) VALUES (?, ?, ?, 0, CURRENT_TIMESTAMP)
                                  ON CONFLICT(pos_employee, pos_job) DO UPDATE SET role_id = excluded.role_id, ignored = 0, updated_at = CURRENT_TIMESTAMP""", (pos_employee, pos_job, role_id))
                else:
                    db.execute("""INSERT INTO pos_employee_job_mappings (pos_employee, pos_job, role_id, ignored, updated_at) VALUES (?, ?, NULL, 1, CURRENT_TIMESTAMP)
                                  ON CONFLICT(pos_employee, pos_job) DO UPDATE SET role_id = NULL, ignored = 1, updated_at = CURRENT_TIMESTAMP""", (pos_employee, pos_job))

        dates = [row["_date"] for row in rows]
        cur = db.execute("INSERT INTO pos_imports (filename, fingerprint, date_min, date_max, row_count) VALUES (?, ?, ?, ?, ?)", (filename, fingerprint, min(dates).isoformat(), max(dates).isoformat(), len(rows)))
        import_id = int(cur.lastrowid)
        employee_names = {int(row["id"]): row["name"] for row in db.execute("SELECT id, name FROM employees")}
        imported_patterns = 0
        skipped_patterns = 0
        for pattern in patterns:
            mapping = pattern_mapping.get((pattern["employee"], pattern["job"]))
            if not mapping:
                skipped_patterns += 1
                continue
            employee_id, role_id = mapping
            role = valid_roles[role_id]
            source = f"POS:{import_id}:{pattern['date'].isoformat()}"
            db.execute("""INSERT INTO shift_history
                          (role_title, role_subtitle, employee_name, day_index, label, source_sheet, count)
                          VALUES (?, ?, ?, ?, ?, ?, 1)
                          ON CONFLICT(role_title, role_subtitle, employee_name, day_index, label, source_sheet)
                          DO UPDATE SET count = count + 1""",
                       (role.title, role.subtitle, employee_names[employee_id].upper(), pattern["day_index"], pattern["label"], source))
            imported_patterns += 1
        db.commit()

    staged_path.unlink(missing_ok=True)
    detail = f"Imported {imported_patterns} learned report-time patterns from {len(rows)} POS entries ({min(dates).isoformat()} through {max(dates).isoformat()})."
    if created_employees or assignments_added:
        detail += f" Created {created_employees} employees and added {assignments_added} role assignments."
    if skipped_patterns:
        detail += f" Skipped {skipped_patterns} patterns whose employee/job mapping was ignored."
    if row_errors:
        detail += f" Ignored {len(row_errors)} invalid CSV rows."
    flash(detail)
    return redirect(url_for("import_history"))


@app.route("/employees")
def manage_employees():
    with closing(get_db()) as db:
        return render_template("employees.html", roles=roles(db), grouped=grouped_employees(db), archived_grouped=grouped_employees(db, archived=True))


@app.route("/api/employees/add", methods=["POST"])
def api_employee_add():
    data = request.get_json(force=True)
    name = (data.get("name") or "").strip().upper()
    role_id = int(data.get("role_id") or 1)
    if not name:
        return jsonify(ok=False, error="Name is required"), 400
    with closing(get_db()) as db:
        max_order = db.execute("SELECT COALESCE(MAX(display_order), 0) FROM employee_assignments WHERE role_id = ? AND archived_at IS NULL", (role_id,)).fetchone()[0]
        cur = db.execute("INSERT INTO employees (name, role_id, active, display_order) VALUES (?, ?, 1, ?)", (name, role_id, int(max_order) + 1))
        assignment = db.execute("INSERT INTO employee_assignments (employee_id, role_id, active, display_order) VALUES (?, ?, 1, ?)", (cur.lastrowid, role_id, int(max_order) + 1))
        db.commit()
    return jsonify(ok=True, id=cur.lastrowid, assignment_id=assignment.lastrowid)


@app.route("/api/employees/<int:employee_id>", methods=["PATCH"])
def api_employee_update(employee_id: int):
    data = request.get_json(force=True)
    fields = []
    params = []
    allowed = {"name": lambda value: str(value).strip().upper()}
    for field, converter in allowed.items():
        if field in data:
            fields.append(f"{field} = ?")
            params.append(converter(data[field]))
    if not fields:
        return jsonify(ok=True)
    params.append(employee_id)
    with closing(get_db()) as db:
        db.execute(f"UPDATE employees SET {', '.join(fields)} WHERE id = ?", params)
        db.commit()
    return jsonify(ok=True)


@app.route("/api/employees/<int:employee_id>/roles", methods=["POST"])
def api_employee_add_role(employee_id: int):
    data = request.get_json(force=True)
    role_id = int(data["role_id"])
    with closing(get_db()) as db:
        existing = db.execute("SELECT id, archived_at FROM employee_assignments WHERE employee_id = ? AND role_id = ?", (employee_id, role_id)).fetchone()
        if existing:
            db.execute("UPDATE employee_assignments SET active = 1, archived_at = NULL WHERE id = ?", (existing["id"],))
            assignment_id = int(existing["id"])
        else:
            max_order = db.execute("SELECT COALESCE(MAX(display_order), 0) FROM employee_assignments WHERE role_id = ? AND archived_at IS NULL", (role_id,)).fetchone()[0]
            cur = db.execute("INSERT INTO employee_assignments (employee_id, role_id, active, display_order) VALUES (?, ?, 1, ?)", (employee_id, role_id, int(max_order) + 1))
            assignment_id = int(cur.lastrowid)
        db.execute("UPDATE employees SET active = 1, archived_at = NULL WHERE id = ?", (employee_id,))
        db.commit()
    return jsonify(ok=True, assignment_id=assignment_id)


@app.route("/api/assignments/<int:assignment_id>/archive", methods=["POST"])
def api_assignment_archive(assignment_id: int):
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    with closing(get_db()) as db:
        row = db.execute("SELECT employee_id FROM employee_assignments WHERE id = ?", (assignment_id,)).fetchone()
        if not row:
            return jsonify(ok=False, error="Assignment not found"), 404
        db.execute("UPDATE employee_assignments SET active = 0, archived_at = ? WHERE id = ?", (now, assignment_id))
        remaining = db.execute("SELECT COUNT(*) FROM employee_assignments WHERE employee_id = ? AND active = 1 AND archived_at IS NULL", (row["employee_id"],)).fetchone()[0]
        if not remaining:
            db.execute("UPDATE employees SET active = 0, archived_at = ? WHERE id = ?", (now, row["employee_id"]))
        db.commit()
    return jsonify(ok=True)


@app.route("/api/assignments/<int:assignment_id>/unarchive", methods=["POST"])
def api_assignment_unarchive(assignment_id: int):
    with closing(get_db()) as db:
        row = db.execute("SELECT employee_id FROM employee_assignments WHERE id = ?", (assignment_id,)).fetchone()
        if not row:
            return jsonify(ok=False, error="Assignment not found"), 404
        db.execute("UPDATE employee_assignments SET active = 1, archived_at = NULL WHERE id = ?", (assignment_id,))
        db.execute("UPDATE employees SET active = 1, archived_at = NULL WHERE id = ?", (row["employee_id"],))
        db.commit()
    return jsonify(ok=True)


@app.route("/api/employees/reorder", methods=["POST"])
def api_employee_reorder():
    data = request.get_json(force=True)
    with closing(get_db()) as db:
        for role_id, assignment_ids in (data.get("roles") or {}).items():
            for index, assignment_id in enumerate(assignment_ids, start=1):
                db.execute("UPDATE employee_assignments SET role_id = ?, display_order = ? WHERE id = ?", (int(role_id), index, int(assignment_id)))
        db.commit()
    return jsonify(ok=True)


@app.route("/backup")
def backup_database():
    init_db()
    with NamedTemporaryFile(delete=False, suffix=".sqlite3") as temp:
        temp_path = Path(temp.name)
    try:
        source = sqlite3.connect(DB_PATH)
        destination = sqlite3.connect(temp_path)
        source.backup(destination)
        destination.close()
        source.close()
        data = temp_path.read_bytes()
    finally:
        temp_path.unlink(missing_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d-%H%M")
    return send_file(io.BytesIO(data), as_attachment=True, download_name=f"gaucho-schedule-backup-{stamp}.sqlite3", mimetype="application/octet-stream")


@app.route("/export/<week_start>.csv")
def export_csv(week_start: str):
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        _, grouped, shifts = grouped_schedule(db, start)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["GAUCHO URBANO EMPLOYEE SCHEDULE", start.isoformat(), "through", (start + timedelta(days=6)).isoformat()])
    for role, employees_for_role in grouped:
        writer.writerow([])
        writer.writerow([role.title, role.subtitle, *[short_date_label(day) for day in week_dates(start)]])
        for employee in employees_for_role:
            writer.writerow([employee.name, "", *[shifts.get((employee.assignment_id, index), "") for index in range(7)]])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename=gaucho-schedule-{start.isoformat()}.csv"})


@app.route("/export/<week_start>.xlsx")
def export_xlsx(week_start: str):
    start = parse_week_start(week_start)
    with closing(get_db()) as db:
        _, grouped, shifts = grouped_schedule(db, start)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = start.strftime("%b %d")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EDEDED")
    row_number = 1
    date_format = "%-m/%-d/%Y" if os.name != "nt" else "%#m/%#d/%Y"
    worksheet.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=9)
    worksheet.cell(row=row_number, column=1, value=f"GAUCHO URBANO EMPLOYEE SCHEDULE    {start.strftime(date_format)} through {(start + timedelta(days=6)).strftime(date_format)}").font = Font(bold=True, size=14)
    row_number += 2
    for role, employees_for_role in grouped:
        worksheet.cell(row=row_number, column=1, value=f"{role.title} {role.subtitle}".strip()).font = Font(bold=True)
        worksheet.cell(row=row_number, column=1).fill = header_fill
        for column, value in enumerate(week_dates(start), start=3):
            cell = worksheet.cell(row=row_number, column=column, value=short_date_label(value))
            cell.font = Font(bold=True, italic=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.fill = header_fill
        for column in range(1, 10):
            worksheet.cell(row=row_number, column=column).border = border
        row_number += 1
        for employee in employees_for_role:
            worksheet.cell(row=row_number, column=1, value=employee.name).font = Font(bold=True)
            for day_index in range(7):
                worksheet.cell(row=row_number, column=day_index + 3, value=shifts.get((employee.assignment_id, day_index), ""))
            for column in range(1, 10):
                worksheet.cell(row=row_number, column=column).border = border
                worksheet.cell(row=row_number, column=column).alignment = Alignment(horizontal="center", wrap_text=True)
            worksheet.cell(row=row_number, column=1).alignment = Alignment(horizontal="left")
            row_number += 1
        row_number += 1
    for column in range(1, 10):
        worksheet.column_dimensions[get_column_letter(column)].width = 14 if column > 1 else 18
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 2
    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)
    return send_file(file_stream, as_attachment=True, download_name=f"gaucho-schedule-{start.isoformat()}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def open_browser() -> None:
    port = int(os.environ.get("PORT", "5000"))
    webbrowser.open(f"http://127.0.0.1:{port}")


if __name__ == "__main__":
    init_db()
    default_host = "0.0.0.0" if os.environ.get("REPL_ID") else "127.0.0.1"
    if IS_PACKAGED and default_host == "127.0.0.1":
        threading.Timer(1.0, open_browser).start()
    debug = os.environ.get("GAUCHO_SCHEDULE_DEBUG", "0") == "1" and not IS_PACKAGED
    app.run(host=os.environ.get("GAUCHO_SCHEDULE_HOST", default_host), port=int(os.environ.get("PORT", "5000")), debug=debug, use_reloader=debug)
